"""
excel_to_neural_json.py
=======================
Converts the ARA Neural input Excel (or CSV) file into TWO output files:

  1. neural_schema.json         — Full detail (backend only, never served to browser).
                                  Contains every record with all fields.
  2. neural_schema_summary.json — Lightweight summary list (frontend, Stage 1).
                                  Contains only per-group metadata; records[] stripped.

Architecture
------------
  Frontend Stage 1  →  GET /api/ara-neural/fullkeys
                        served from neural_schema_summary.json
                        (or real backend endpoint in production)

  Frontend Stage 2  →  GET /api/ara-neural/fullkeys/{groupId}
                        served from neural_schema.json lookup
                        (or real backend endpoint in production)
                        Called ONLY when the user selects a FULL_KEY group.

Usage
-----
    # From CSV (quick test):
    python scripts/excel_to_neural_json.py --input input/ara_neural_input.csv

    # From Excel:
    python scripts/excel_to_neural_json.py --input input/ara_neural_input.xlsx

    # Specify output paths explicitly:
    python scripts/excel_to_neural_json.py \
        --input          input/ara_neural_input.xlsx \
        --output         ara-workspace/projects/arg-portal/src/assets/data/neural_schema.json \
        --summary-output ara-workspace/projects/arg-portal/src/assets/data/neural_schema_summary.json

Requirements
------------
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLUMNS: list[str] = [
    "FULL_KEY",
    "BANK_ACCOUNT",
    "RECON_SYSTEM_PLATFORM",
    "RECON_SYSTEM_DATABASE",
    "RECON_BALANCE_POOL",
    "RECON_ACCOUNT",
    "ACCOUNT_STATUS",
    "ACCOUNT_OWNER_SOEID",
    "PROOF_OWNER_SOEID",
    "BANK_ACCOUNT_TYPE",
    "REGION",
    "COUNTRY",
    "AO_SOE_ID",
    "AO_NAME",
    "LINE_OF_BUSINESS",
    "BSS_ACCOUNT_TYPE",
    "BSER_REPORTABLE",
    "RISK_TYPE",
    "REVIEW_STATUS",
    "DDQ_STATUS",
    "ARG_REVIEW_OWNER",
    "PO_SOEID",
    "PO_NAME",
    "AO_STATUS",
]

OUTPUT_DEFAULT = os.path.join(
    "ara-workspace", "projects", "arg-portal",
    "src", "assets", "data", "neural_schema.json"
)

SUMMARY_OUTPUT_DEFAULT = os.path.join(
    "ara-workspace", "projects", "arg-portal",
    "src", "assets", "data", "neural_schema_summary.json"
)


# ─────────────────────────────────────────────────────────────────────────────
# Excel / CSV helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_dataframe(input_path: Path) -> pd.DataFrame:
    """Read Excel or CSV into a DataFrame. Returns normalised column names."""
    suffix = input_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(input_path, sheet_name=0, dtype=str)
    elif suffix == ".csv":
        df = pd.read_csv(input_path, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Use .xlsx, .xls, or .csv")

    # Strip surrounding whitespace from column names and values
    df.columns = [c.strip().upper() for c in df.columns]
    # pandas >= 2.1: applymap was renamed to map
    _map_fn = df.map if hasattr(df, "map") else df.applymap
    df = _map_fn(lambda v: v.strip() if isinstance(v, str) else v)

    return df


def validate_columns(df: pd.DataFrame) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Input file is missing required columns: {missing}\n"
            f"Columns found: {list(df.columns)}"
        )


def safe_str(val: Any, fallback: str = "") -> str:
    """Return a clean string or the fallback for NaN / None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return fallback
    s = str(val).strip()
    return s if s.lower() not in ("nan", "none", "") else fallback


# ─────────────────────────────────────────────────────────────────────────────
# Record builder
# ─────────────────────────────────────────────────────────────────────────────

def build_record(row: pd.Series, record_index: int) -> dict:
    """Map a single DataFrame row into the nested record schema."""
    def s(col: str, fb: str = "") -> str:
        return safe_str(row.get(col), fb)

    record_id = f"rec_{record_index:06d}"

    return {
        "recordId": record_id,
        "central": {
            "bankAccount":   s("BANK_ACCOUNT"),
            "accountStatus": s("ACCOUNT_STATUS"),
            "accountType":   s("BANK_ACCOUNT_TYPE"),
            "region":        s("REGION"),
            "country":       s("COUNTRY"),
            "lineOfBusiness": s("LINE_OF_BUSINESS"),
            "riskType":      s("RISK_TYPE"),
        },
        "system": {
            "platform":     s("RECON_SYSTEM_PLATFORM"),
            "database":     s("RECON_SYSTEM_DATABASE"),
            "balancePool":  s("RECON_BALANCE_POOL"),
            "reconAccount": s("RECON_ACCOUNT"),
        },
        "ownership": {
            "accountOwner": {
                "soeid": s("ACCOUNT_OWNER_SOEID"),
            },
            "proofOwner": {
                "soeid": s("PROOF_OWNER_SOEID"),
            },
            "argReviewOwner": {
                "soeid": s("ARG_REVIEW_OWNER"),
            },
        },
        "approval": {
            "ao": {
                "soeid":  s("AO_SOE_ID"),
                "name":   s("AO_NAME"),
                "status": s("AO_STATUS"),
            },
            "po": {
                "soeid": s("PO_SOEID"),
                "name":  s("PO_NAME"),
            },
            "reviewStatus": s("REVIEW_STATUS"),
            "ddqStatus":    s("DDQ_STATUS"),
        },
        "usage": {
            "bssAccountType": s("BSS_ACCOUNT_TYPE"),
            "bserReportable": s("BSER_REPORTABLE"),
        },
        "raw": {
            "fullKey":              s("FULL_KEY"),
            "bankAccount":          s("BANK_ACCOUNT"),
            "reconSystemPlatform":  s("RECON_SYSTEM_PLATFORM"),
            "reconSystemDatabase":  s("RECON_SYSTEM_DATABASE"),
            "reconBalancePool":     s("RECON_BALANCE_POOL"),
            "reconAccount":         s("RECON_ACCOUNT"),
            "accountStatus":        s("ACCOUNT_STATUS"),
            "accountOwnerSoeid":    s("ACCOUNT_OWNER_SOEID"),
            "proofOwnerSoeid":      s("PROOF_OWNER_SOEID"),
            "bankAccountType":      s("BANK_ACCOUNT_TYPE"),
            "region":               s("REGION"),
            "country":              s("COUNTRY"),
            "aoSoeId":              s("AO_SOE_ID"),
            "aoName":               s("AO_NAME"),
            "lineOfBusiness":       s("LINE_OF_BUSINESS"),
            "bssAccountType":       s("BSS_ACCOUNT_TYPE"),
            "bserReportable":       s("BSER_REPORTABLE"),
            "riskType":             s("RISK_TYPE"),
            "reviewStatus":         s("REVIEW_STATUS"),
            "ddqStatus":            s("DDQ_STATUS"),
            "argReviewOwner":       s("ARG_REVIEW_OWNER"),
            "poSoeid":              s("PO_SOEID"),
            "poName":               s("PO_NAME"),
            "aoStatus":             s("AO_STATUS"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Group builder
# ─────────────────────────────────────────────────────────────────────────────

def _unique_nonempty(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for v in values:
        if v and v not in seen:
            seen.add(v)
            result.append(v)
    return result


def build_group(full_key: str, group_df: pd.DataFrame,
                group_index: int, record_counter: list[int]) -> dict:
    """Build one group object containing all records for a single FULL_KEY."""

    records: list[dict] = []
    for _, row in group_df.iterrows():
        record_counter[0] += 1
        records.append(build_record(row, record_counter[0]))

    def col_values(col: str) -> list[str]:
        return [safe_str(v) for v in group_df[col].tolist()]

    # Derive primary region from the most common value in the group
    region_series = group_df["REGION"].dropna()
    primary_region = region_series.mode()[0] if not region_series.empty else ""

    summary = {
        "region":            primary_region,
        "recordCount":       len(records),
        "countryCount":      len(_unique_nonempty(col_values("COUNTRY"))),
        "bankAccountCount":  len(_unique_nonempty(col_values("BANK_ACCOUNT"))),
        "accountStatuses":   _unique_nonempty(col_values("ACCOUNT_STATUS")),
        "platforms":         _unique_nonempty(col_values("RECON_SYSTEM_PLATFORM")),
        "databases":         _unique_nonempty(col_values("RECON_SYSTEM_DATABASE")),
        "aoNames":           _unique_nonempty(col_values("AO_NAME")),
        "poNames":           _unique_nonempty(col_values("PO_NAME")),
        "reviewStatuses":    _unique_nonempty(col_values("REVIEW_STATUS")),
        "ddqStatuses":       _unique_nonempty(col_values("DDQ_STATUS")),
    }

    return {
        "groupId":      f"fk_{group_index:06d}",
        "fullKey":      full_key,
        "displayTitle": full_key,
        "records":      records,
        "summary":      summary,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Summary builder  (strips records[] — frontend Stage 1 payload)
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_item(group: dict) -> dict:
    """Return a lightweight summary dict from a full group — no records[] array."""
    s = group["summary"]
    return {
        "groupId":       group["groupId"],
        "fullKey":       group["fullKey"],
        "region":        s["region"],
        "recordCount":   s["recordCount"],
        "countryCount":  s["countryCount"],
        "reviewStatuses": s["reviewStatuses"],
        "platforms":     s["platforms"],
    }


def write_summary(groups: list[dict], summary_path: Path, generated_at: str) -> None:
    """
    Write neural_schema_summary.json — the frontend Stage 1 payload.

    This file is safe to serve to the browser: it contains only group
    metadata.  records[] (which may be thousands of rows per group) is
    intentionally excluded.

    In production this is replaced by a real API endpoint:
        GET /api/ara-neural/fullkeys
    """
    items = [build_summary_item(g) for g in groups]
    summary: dict = {
        "version":     "1.0",
        "generatedAt": generated_at,
        "total":       len(items),
        "page":        1,
        "pageSize":    len(items),
        "items":       items,
    }
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with open(summary_path, "w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False)
    print(f"[+] Summary written: {summary_path}  ({len(items)} items, no records[])")


# ─────────────────────────────────────────────────────────────────────────────
# Main conversion
# ─────────────────────────────────────────────────────────────────────────────

def convert(input_path: Path, output_path: Path, summary_path: Path) -> None:
    print(f"[+] Loading input  : {input_path}")
    df = load_dataframe(input_path)

    print(f"[+] Validating columns …")
    validate_columns(df)

    print(f"[+] Rows loaded    : {len(df)}")

    groups: list[dict] = []
    record_counter = [0]          # mutable counter shared across groups

    for group_index, (full_key, group_df) in enumerate(df.groupby("FULL_KEY", sort=False), start=1):
        group = build_group(str(full_key), group_df, group_index, record_counter)
        groups.append(group)

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # ── Full detail JSON (backend / mock only — never served to browser) ──────
    full_output: dict = {
        "version":     "1.0",
        "generatedAt": generated_at,
        "groups":      groups,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(full_output, fh, indent=2, ensure_ascii=False)

    total_records = sum(len(g["records"]) for g in groups)
    print(f"[+] Groups written : {len(groups)}")
    print(f"[+] Records total  : {total_records}")
    print(f"[+] Full JSON      : {output_path}")

    # ── Summary JSON (frontend Stage 1 — records[] stripped) ─────────────────
    write_summary(groups, summary_path, generated_at)


# ─────────────────────────────────────────────────────────────────────────────
# Excel template generator (bonus helper)
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel_from_csv(csv_path: Path, xlsx_path: Path) -> None:
    """Read the CSV and write it out as a properly formatted Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[!] openpyxl not installed — skipping Excel generation. "
              "Run: pip install openpyxl")
        return

    df = pd.read_csv(csv_path, dtype=str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARA Neural Input"

    # Header style
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 2, 14)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Alternate row fills
    fill_even = PatternFill("solid", fgColor="EBF0FA")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10)
    data_align = Alignment(vertical="center")

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value="" if pd.isna(value) else str(value))
            cell.fill  = fill
            cell.font  = data_font
            cell.alignment = data_align
            cell.border = thin_border

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"[+] Excel template  : {xlsx_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert ARA Neural Excel/CSV input into TWO output files:\n"
            "  1. neural_schema.json         — full detail (backend / mock only)\n"
            "  2. neural_schema_summary.json — lightweight summaries (frontend Stage 1)"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to input .xlsx, .xls, or .csv file.",
    )
    parser.add_argument(
        "--output", "-o",
        default=OUTPUT_DEFAULT,
        help=f"Path to write neural_schema.json (full detail). Default: {OUTPUT_DEFAULT}",
    )
    parser.add_argument(
        "--summary-output", "-s",
        default=SUMMARY_OUTPUT_DEFAULT,
        help=(
            f"Path to write neural_schema_summary.json (frontend Stage 1, no records[]). "
            f"Default: {SUMMARY_OUTPUT_DEFAULT}"
        ),
    )
    parser.add_argument(
        "--generate-excel",
        action="store_true",
        help="Also generate a formatted .xlsx template from the CSV input.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_path   = Path(args.input)
    output_path  = Path(args.output)
    summary_path = Path(args.summary_output)

    if not input_path.exists():
        print(f"[ERROR] Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Optionally generate Excel from CSV first
    if args.generate_excel:
        xlsx_path = input_path.with_suffix(".xlsx")
        print(f"[+] Generating Excel: {xlsx_path}")
        generate_excel_from_csv(input_path, xlsx_path)

    # Run conversion — emits both full JSON and summary JSON
    convert(input_path, output_path, summary_path)
    print("[✓] Done.")


if __name__ == "__main__":
    main()
